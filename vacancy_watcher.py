# -*- coding: utf-8 -*-
"""Сборщик вакансий с нескольких сайтов: парсинг -> SQLite -> новые в Telegram.

  python vacancy_watcher.py collect          собрать вакансии, записать новые в базу
  python vacancy_watcher.py send             отправить в Telegram то, что ещё не уходило
  python vacancy_watcher.py run              собрать и сразу отправить
  python vacancy_watcher.py stats            что лежит в базе
  python vacancy_watcher.py export           выгрузить всё в Excel

Ключи: --pages N (сколько страниц каталога брать), --dry (не отправлять, только показать).
Токен и чат берутся из переменных окружения TG_TOKEN и TG_CHAT.
"""
import argparse
import html
import os
import re
import sqlite3
import sys
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import quote, urljoin

import requests
from bs4 import BeautifulSoup

HERE = Path(__file__).parent
DB_PATH = HERE / "vacancies.db"


def load_env():
    """Читает .env рядом со скриптом. На сервере так удобнее, чем export."""
    path = HERE / ".env"
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, value = line.partition("=")
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


load_env()

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/127.0.0.0 Safari/537.36")
HEADERS = {
    "User-Agent": UA,
    "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

# --- фильтр: что и где ищем ---------------------------------------------
# Москва: на hh это area=1, на rabota.ru geo_id=4400.
HH_AREA = 1
RABOTA_GEO = 4400

# Профессиональные роли hh, из которых складывается «айти»:
# 96 разработчик, 10 аналитик, 104 руководитель разработки, 113 сисадмин,
# 124 тестировщик, 156 BI и аналитик данных, 160 DevOps, 165 Data Science.
HH_IT_ROLES = [96, 10, 104, 113, 124, 156, 160, 165]
HH_ROLES_PARAM = "&".join(f"professional_role={r}" for r in HH_IT_ROLES)

# У rabota.ru рубрика ИТ дорисовывается скриптами и обычным запросом не
# берётся, зато работает поиск по слову вместе с городом. Поэтому айти
# там собирается набором запросов.
IT_QUERIES = [
    "разработчик", "программист", "аналитик", "тестировщик",
    "devops", "системный администратор", "python", "1С",
]

# Поиск по слову тянет лишнее: по «аналитик» приходит химик-аналитик,
# по «1С» — оператор 1С в строительном магазине. Поэтому там, где сайт
# ищет текстом, название проверяется дополнительно.
NOT_IT_RE = re.compile(
    r"химик|повар|официант|бариста|водител|продав|кассир|прораб|грузчик|швея|"
    r"парикмахер|уборщ|курьер|охранник|кладовщик|комплектовщик|мерчендайзер|"
    r"менеджер по продажам|торговый представитель|сварщик|электромонтаж|"
    r"оператор 1с|операционист|преподавател|репетитор|товарному учет|"
    r"сопровождению сделок|на склад", re.I)
IT_RE = re.compile(
    r"разработ|программист|developer|devops|аналитик|analyst|тестировщ|"
    r"тестирован|\bqa\b|\bsre\b|\bml\b|data|дата|администратор|инженер|"
    r"архитектор|верстальщ|frontend|backend|fullstack|full-stack|"
    r"\bит\b|\bit\b|1c|1с|python|java|php|golang|техподдержк|"
    r"технической поддержки|информационн", re.I)


def looks_like_it(title):
    if NOT_IT_RE.search(title):
        return False
    return bool(IT_RE.search(title))


# Каждый сайт — просто набор селекторов. Новый источник добавляется сюда,
# трогать остальной код не нужно.
SITES = {
    # hh.ru: классы в разметке хэшированные и меняются с релизами, поэтому
    # цепляемся за data-qa, а зарплату берём регексом по тексту карточки.
    # Выборку сужает сама ссылка: город плюс профессиональные роли.
    "hh.ru": {
        "page_url": (f"https://hh.ru/search/vacancy?area={HH_AREA}"
                     f"&{HH_ROLES_PARAM}&page={{page}}"),
        "page_start": 0,
        "card": '[data-qa="vacancy-serp__vacancy"]',
        "title": '[data-qa="serp-item__title"]',
        "company": '[data-qa="vacancy-serp__vacancy-employer"]',
        "salary": None,
        "salary_regex": True,
        "pause": 4.0,
    },
    # superjob.ru отдаёт капчу на любой запрос, включая главную и robots.txt.
    # Селекторы оставлены рабочие: заработает, как только появится доступ
    # (официальный ключ api.superjob.ru или обход капчи).
    "superjob.ru": {
        "page_url": ("https://moscow.superjob.ru/vacancy/search/"
                     "?keywords={query}&page={page}"),
        "queries": IT_QUERIES,
        "filter_it": True,
        "card": "div.f-test-vacancy-item",
        "title": "a.f-test-link-Vakansiya",
        "company": "span.f-test-text-vacancy-item-company-name",
        "salary": None,
        "salary_regex": True,
        "pause": 3.0,
    },
    "career.habr.com": {
        "page_url": "https://career.habr.com/vacancies?type=all&page={page}",
        "card": "div.vacancy-card",
        "title": "a.vacancy-card__title-link",
        "company": "div.vacancy-card__company a.link-comp",
        "salary": "div.vacancy-card__salary",
    },
    "geekjob.ru": {
        "page_url": "https://geekjob.ru/vacancies?page={page}",
        "card": "li.collection-item.avatar",
        "title": "p.vacancy-name a.title",
        "company": "p.company-name a",
        "salary": "span.salary",
    },
    "rabota.ru": {
        "page_url": (f"https://www.rabota.ru/vacancy?query={{query}}"
                     f"&geo_id={RABOTA_GEO}&page={{page}}"),
        "queries": IT_QUERIES,
        "filter_it": True,
        "card": "div.vacancy-preview-card__wrapper",
        "title": "h3.vacancy-preview-card__title a",
        "company": "span.vacancy-preview-card__company-name",
        "salary": "div.vacancy-preview-card__salary",
    },
}

NO_SALARY = re.compile(r"не указана|по договор|договорённост|договоренност", re.I)
# «300 000 – 380 000 ₽ за месяц», «от 55 000 руб.», «до 8000 €»
SALARY_RE = re.compile(
    r"(?:от|до)?\s*\d[\d  ]{2,}(?:\s*[–—-]\s*\d[\d  ]{2,})?\s*(?:₽|руб|€|\$)[^,.;]{0,20}",
    re.I)


def clean(text):
    # rabota.ru отдаёт названия с двойным экранированием: АО &quot;Аметист&quot;
    text = html.unescape(html.unescape(text or ""))
    return re.sub(r"\s+", " ", text.replace("\xa0", " ")).strip()


def norm_salary(text):
    text = clean(text)
    if not text or NO_SALARY.search(text):
        return "не указана"
    return text


# --- база ---------------------------------------------------------------

def db_connect():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS vacancies (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            site       TEXT NOT NULL,
            title      TEXT NOT NULL,
            company    TEXT,
            salary     TEXT,
            url        TEXT NOT NULL UNIQUE,
            fingerprint TEXT NOT NULL,
            found_at   TEXT NOT NULL,
            sent_at    TEXT
        )
    """)
    # по отпечатку ищем клонов: та же вакансия, перевыложенная под новой ссылкой
    conn.execute("CREATE INDEX IF NOT EXISTS idx_fp ON vacancies(fingerprint)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_sent ON vacancies(sent_at)")
    conn.commit()
    return conn


def fingerprint(site, title, company, salary):
    return "|".join([site, title.lower(), (company or "").lower(), salary])


# --- парсинг ------------------------------------------------------------

def pick(card, selector):
    if not selector:
        return ""
    node = card.select_one(selector)
    return clean(node.get_text(" ", strip=True)) if node else ""


def find_salary(card, cfg):
    """Зарплата либо по своему селектору, либо поиском по тексту карточки."""
    if cfg.get("salary"):
        return norm_salary(pick(card, cfg["salary"]))
    if cfg.get("salary_regex"):
        match = SALARY_RE.search(clean(card.get_text(" ", strip=True)))
        if match:
            return clean(match.group(0))
    return "не указана"


def parse_page(site, cfg, page, query=""):
    url = cfg["page_url"].format(page=page, query=quote(query))
    resp = requests.get(url, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    title_tag = soup.title.get_text(strip=True).lower() if soup.title else ""
    if "captcha" in title_tag or "капча" in title_tag:
        raise RuntimeError("сайт отдал капчу вместо страницы")

    rows = []
    for card in soup.select(cfg["card"]):
        title_node = card.select_one(cfg["title"])
        if not title_node:
            continue
        title = clean(title_node.get_text(" ", strip=True))
        if not title:
            continue
        # на hh выборку уже сузили роли в ссылке, здесь чистим текстовый поиск
        if cfg.get("filter_it") and not looks_like_it(title):
            continue

        rows.append({
            "site": site,
            "title": title,
            "company": pick(card, cfg.get("company")) or "не указана",
            "salary": find_salary(card, cfg),
            # хвост запроса режем: на hh и rabota.ru там метки поиска,
            # из-за которых одна вакансия каждый раз выглядит как новая
            "url": urljoin(url, title_node.get("href", "")).split("?")[0],
        })
    return rows


def collect(pages, only=None):
    conn = db_connect()
    now = datetime.now().isoformat(timespec="seconds")
    report = {}

    targets = {s: c for s, c in SITES.items() if not only or s in only}
    for site, cfg in targets.items():
        added = dup_url = dup_clone = 0
        start = cfg.get("page_start", 1)
        # сайт либо ищет по набору запросов (rabota, superjob),
        # либо отдаёт готовую выборку по фильтру прямо в ссылке (hh)
        queries = cfg.get("queries") or [""]

        for query in queries:
            for page in range(start, start + pages):
                try:
                    rows = parse_page(site, cfg, page, query=query)
                except Exception as e:
                    label = f"«{query}» " if query else ""
                    print(f"  {site} {label}стр.{page}: не открылась ({e})")
                    break
                if not rows:
                    break

                for r in rows:
                    fp = fingerprint(site, r["title"], r["company"], r["salary"])

                    cur = conn.execute("SELECT 1 FROM vacancies WHERE url = ?", (r["url"],))
                    if cur.fetchone():
                        dup_url += 1
                        continue

                    cur = conn.execute("SELECT 1 FROM vacancies WHERE fingerprint = ?", (fp,))
                    if cur.fetchone():
                        dup_clone += 1
                        continue

                    conn.execute(
                        "INSERT INTO vacancies (site, title, company, salary, url,"
                        " fingerprint, found_at) VALUES (?,?,?,?,?,?,?)",
                        (site, r["title"], r["company"], r["salary"], r["url"], fp, now),
                    )
                    added += 1
                conn.commit()
                time.sleep(cfg.get("pause", 1.5))

        report[site] = (added, dup_url, dup_clone)
        print(f"{site}: новых {added}, уже было {dup_url}, клонов отсеяно {dup_clone}")

    conn.close()
    return report


# --- отправка -----------------------------------------------------------

def escape_html(text):
    return (text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))


def build_messages(rows, chunk=10):
    """Собирает сообщения для Telegram: по 10 вакансий, чтобы не упереться в лимит."""
    messages = []
    for start in range(0, len(rows), chunk):
        part = rows[start:start + chunk]
        lines = [f"<b>Новые вакансии: {len(rows)}</b>" if start == 0 else "<b>продолжение</b>"]
        for r in part:
            lines.append(
                f"\n<a href=\"{r['url']}\">{escape_html(r['title'])}</a>\n"
                f"{escape_html(r['company'])} · {escape_html(r['salary'])}\n"
                f"<i>{r['site']}</i>"
            )
        messages.append("\n".join(lines))
    return messages


def send(dry=False, limit=None, keep=False):
    conn = db_connect()
    cur = conn.execute(
        "SELECT id, site, title, company, salary, url FROM vacancies"
        " WHERE sent_at IS NULL ORDER BY site, id"
        + (f" LIMIT {int(limit)}" if limit else "")
    )
    rows = [dict(zip(("id", "site", "title", "company", "salary", "url"), r))
            for r in cur.fetchall()]

    if not rows:
        print("новых вакансий нет, отправлять нечего")
        conn.close()
        return 0

    messages = build_messages(rows)
    print(f"к отправке: {len(rows)} вакансий в {len(messages)} сообщениях")

    if dry:
        print("\n--- как это уйдёт в Telegram (первое сообщение) ---")
        print(re.sub(r"<[^>]+>", "", messages[0]))
        conn.close()
        return 0

    token, chat = os.environ.get("TG_TOKEN"), os.environ.get("TG_CHAT")
    if not token or not chat:
        print("нет TG_TOKEN / TG_CHAT в переменных окружения")
        conn.close()
        return 1

    api = f"https://api.telegram.org/bot{token}/sendMessage"
    for i, text in enumerate(messages, 1):
        resp = requests.post(api, timeout=30, data={
            "chat_id": chat,
            "text": text,
            "parse_mode": "HTML",
            "disable_web_page_preview": True,
        })
        ok = resp.json().get("ok")
        print(f"  сообщение {i}/{len(messages)}: {'ушло' if ok else resp.text[:200]}")
        if not ok:
            conn.close()
            return 1
        time.sleep(1)

    if keep:
        conn.close()
        print("отправлено, но в базе вакансии остались новыми (показ, не рассылка)")
        return 0

    now = datetime.now().isoformat(timespec="seconds")
    conn.executemany("UPDATE vacancies SET sent_at = ? WHERE id = ?",
                     [(now, r["id"]) for r in rows])
    conn.commit()
    conn.close()
    print("отправлено, вакансии помечены как отправленные")
    return 0


# --- прочее -------------------------------------------------------------

def stats():
    conn = db_connect()
    total = conn.execute("SELECT COUNT(*) FROM vacancies").fetchone()[0]
    unsent = conn.execute(
        "SELECT COUNT(*) FROM vacancies WHERE sent_at IS NULL").fetchone()[0]
    print(f"всего в базе: {total}, ещё не отправлено: {unsent}\n")
    print(f"{'сайт':<20} {'вакансий':>9} {'с зарплатой':>12}")
    for site, cnt, with_salary in conn.execute(
        "SELECT site, COUNT(*), SUM(salary != 'не указана')"
        " FROM vacancies GROUP BY site ORDER BY COUNT(*) DESC"
    ):
        print(f"{site:<20} {cnt:>9} {with_salary or 0:>12}")
    conn.close()


def export():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    conn = db_connect()
    rows = conn.execute(
        "SELECT site, title, company, salary, url, found_at FROM vacancies"
        " ORDER BY site, id").fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Вакансии"
    headers = ["№", "Сайт", "Название вакансии", "Компания", "Зарплата", "Ссылка", "Найдена"]
    ws.append(headers)

    fill = PatternFill("solid", fgColor="1F3A5F")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")

    for i, (site, title, company, salary, url, found) in enumerate(rows, 1):
        ws.append([i, site, title, company, salary, url, found.replace("T", " ")])
        cell = ws.cell(row=i + 1, column=6)
        cell.hyperlink = url
        cell.font = Font(color="0563C1", underline="single")

    for col, w in enumerate([5, 18, 50, 26, 22, 58, 19], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{len(rows) + 1}"

    out = HERE / "вакансии_3_сайта.xlsx"
    wb.save(out)
    print(f"выгружено {len(rows)} строк: {out}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("cmd", choices=["collect", "send", "run", "stats", "export"])
    ap.add_argument("--pages", type=int, default=2)
    ap.add_argument("--dry", action="store_true")
    ap.add_argument("--limit", type=int, help="отправить не больше N вакансий")
    ap.add_argument("--keep", action="store_true",
                    help="отправить, но не помечать отправленными (показать образец)")
    ap.add_argument("--sites", help="только эти сайты, через запятую")
    ap.add_argument("--db", help="файл базы (по умолчанию vacancies.db)")
    args = ap.parse_args()

    if args.db:
        global DB_PATH
        DB_PATH = HERE / args.db

    only = [s.strip() for s in args.sites.split(",")] if args.sites else None
    if args.cmd in ("collect", "run"):
        collect(args.pages, only=only)
    if args.cmd in ("send", "run"):
        return send(dry=args.dry, limit=args.limit, keep=args.keep)
    if args.cmd == "stats":
        stats()
    if args.cmd == "export":
        export()
    return 0


if __name__ == "__main__":
    sys.exit(main())
